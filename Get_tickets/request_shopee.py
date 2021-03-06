"""
目標
1. 每天自動執行
2. 產生統計報表
3. 去除與關鍵字無關的商品資訊: regular expression, regex
"""


import requests
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import const


headers = {
    "referer": "https://shopee.tw/search?keyword=google%20pixel",
    "Cookie": const.Cookie,
    "User-Agent": const.User_Agent,
    'if-none-match-': const.if_none_match_
}


def retrieve_shopee_item(keyword, accurate, data_count, search_mode):

    file_name = "price_trace.xlsx"

    if search_mode == 1:
        # new excel file
        wb = Workbook()
    elif search_mode == 2:
        # old excel file
        wb = load_workbook(file_name)

    ws = wb.active
    ws["A1"] = "更新日期"
    ws["B1"] = "商品名稱"
    ws["C1"] = "price_max"
    ws["D1"] = "price_min"
    ws["E1"] = "price"
    ws["F1"] = "Is_Accu"

    url = f"https://shopee.tw/api/v2/search_items/?by=relevancy&keyword={keyword}&limit={data_count}&newest=0&order=desc&page_type=search"

    res = requests.get(url,
                       headers=headers,
                       allow_redirects=False)
    json_file = res.json()

    # all items found by keyword
    total_item = json_file.get("items")

    for items in total_item:
        item_name = items.get("name")

        # remove the last 5 numbers
        max_price = str(items.get("price_max"))[:-5]
        min_price = str(items.get("price_min"))[:-5]
        price = str(items.get("price"))[:-5]

        ws.append([datetime.date.today(), item_name,
                   max_price, min_price, price, "N"])

    for items in total_item:
        item_name = items.get("name")
        if accurate in item_name:
            max_price = str(items.get("price_max"))[:-5]
            min_price = str(items.get("price_min"))[:-5]
            price = str(items.get("price"))[:-5]
            ws.append([datetime.date.today(), item_name,
                       max_price, min_price, price, "Y"])

    wb.save(file_name)


if __name__ == "__main__":
    search_mode = int(
        input("將搜尋符合關鍵字的蝦皮商品，請選擇要新開檔案或沿用舊檔，新檔案輸入1，舊檔案輸入2\n"))
    keyword = input("請輸入關鍵字\n")
    accurate = input("必須關鍵字?\n")

    # can only accept 1~100 by now
    data_count = 100
    retrieve_shopee_item(keyword, accurate, data_count, search_mode)
